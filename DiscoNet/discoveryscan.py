#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
discoveryscan module
"""


import paramiko
import threading
from multiprocessing import Process, Lock, JoinableQueue, cpu_count
from openpyxl import Workbook, load_workbook
import ipaddress, time
from datetime import datetime


class _DiscoveryWorkbook:

    def __init__(self, name):
        self.l = Lock()
        self.fname = name
        self.init = False

    def new_sheet(self, sname, rows):
        self.l.acquire()
        w = load_workbook(self.fname) if self.init else Workbook(write_only = True)
        ws = w.create_sheet(sname)
        for row in rows:
            ws.append(row)
        w.save(self.fname)
        self.init = True
        self.l.release()


class DiscoveryScan:
    """Return ``DiscoveryScan``
    
    :param str workbook: path to output xlsx file
    :param str subnets: comma delimited list of networks and IP addresses to scan
    :param str username: SSH username
    :param str password: SSH password
    :param str commands: string listing commands to run, one per line
    :return: an initialized DiscoveryScan object
    :rtype: DiscoveryScan
    
    """

    def __init__(self, workbook, subnets, username, password, commands):


        self.t = datetime.utcnow()
        self.u = username
        self.p = password
        self.c = commands
        self.q = JoinableQueue()
        self.m = 0

        for net in subnets.split(','):
            net = ipaddress.ip_network(net)
            if net.num_addresses > 1:
                for host in net.hosts():
                    self.q.put(str(host))
                    self.m += 1
            else:
                self.q.put(str(net.network_address))
                self.m += 1

        if (2 * cpu_count()) < self.m:
            self.m = (2 * cpu_count())

        ws = []
        ws.append(['Timestamp:', str(datetime.utcnow())])
        ws.append(['Networks:', subnets])
        ws.append(['Username:', self.u])
        ws.append(['Password:', self.p])
        ws.append(['Commands:'])
        x = 1
        for line in self.c.splitlines():
            ws.append([x, line.rstrip()])
            x += 1
        self.wb = _DiscoveryWorkbook(workbook)
        self.wb.new_sheet('Discovery', ws)


    def _run(self):
        if self.q.empty():
            return

        handle = paramiko.SSHClient()
        handle.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        while True:
            host = self.q.get()
            if host is None:
                self.q.task_done()
                return

            asa = False
            try:
                handle.connect(host, username=self.u, password=self.p,
                               look_for_keys=False, timeout=2)
                asa_str = "Type help or '?' for a list of available commands."
                if str(handle.invoke_shell().recv(65535)).find(asa_str) != -1:
                    asa = True
                handle.close()
            except:
                self.q.task_done()
                continue


            x = 1
            for command in self.c.splitlines():
                try:
                    if asa:
                        command = "term pager 0\n" + command + "\nexit\n"
                    else:
                        command = command + "\n"
                    handle.connect(host, username=self.u, password=self.p,
                                   look_for_keys=False, timeout=2)
                    stdin, stdout, stderr = handle.exec_command(command)
                except:
                    continue
                ws = []
                ws.append([command.rstrip()])
                try:
                    stdout.channel.recv_exit_status()
                    for line in stdout:
                        ws.append([line.rstrip()])
                except:
                    continue
                self.wb.new_sheet("%s-%s" % (host, str(x)), ws)
                try:
                    stdin.close()
                    stdout.close()
                    stderr.close()
                    handle.close()
                except:
                    x += 1
            self.q.task_done()

    def _start(self, cb):
        if self.q.empty():
            return

        processes = [Process(target=self._run) for i in range(self.m)]

        try:
            for p in processes:
                p.start()
        finally:
            self.q.join()
            x = 0
            for p in processes:
                if p.is_alive():
                    self.q.put(None)
            for p in processes:
                if p.is_alive():
                    p.join()
            if cb:
                cb()

    def start(self, cb=None):
        """Start Discovery
        
        Starts the discovery process and blocks until complete. Optionally, a callback can
        be specified to make this process non-blocking and the callback will be called
        when complete.
        
        :param function cb: callback method. Defaults to None.
        :return: nothing
        :rtype: None
        """
        if cb:
            thread = threading.Thread(target=self._start, args=(cb,))#, daemon=True)
            thread.start()
        else:
            self._start(cb)

def _main():
    from sys import argv, exit

    usage = ("usage: %s workbook subnets username password commands\n"
             "\n"
             "	workbook	Path to output xlsx file\n"
             "	subnets		Comma delimited list of networks and IP addresses to scan\n"
             "	username	SSH username\n"
             "	password	SSH password\n"
             "	commands	String listing commands to run, one per line\n" % argv[0])

    if len(argv) != 6:
        print(usage)
    else:
        try:
            f = open(argv[1], 'w')
            f.close()
        except:
            print(usage)
            print("Error: Could not open %s for writing" % argv[1])
        try:
            for net in argv[2].split(','):
                net = ipaddress.ip_network(net)
        except:
            print(usage)
            print("Error: subnets must be a comma delimited list of networks and IP "
                  "addresses, e.g.,\n'10.0.0.0/16,192.168.0.0/255.255.255.0,10.10.10.10'")
            exit(1)
        d = DiscoveryScan(argv[1], argv[2], argv[3], argv[4], argv[5])
        d.start()

if __name__ == '__main__':
    _main()
